import re
import smtplib
import string
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path
from pprint import pprint

from bs4 import BeautifulSoup as bs
from datetime import datetime
import requests
import values
import openpyxl
import logging

#
s = requests.session()
homepage = s.get(values.URL)
soup = bs(homepage.text, "html.parser")
csrf_name = soup.find(attrs={"name":values.KEY_CSRF_NAME}).get("value")
csrf_token = soup.find(attrs={"name":values.KEY_CSRF_TOKEN}).get("value")

#print(tag_inputs)
logging.info(f"{values.KEY_CSRF_NAME}: {csrf_name}")
logging.info(f"{values.KEY_CSRF_TOKEN}: {csrf_token}")


login_payload = {
    values.KEY_LOGIN: values.LOGIN,
    values.KEY_PASSWORD: values.PASSWORD,
    values.KEY_CSRF_NAME: csrf_name,
    values.KEY_CSRF_TOKEN: csrf_token
}

login_result = s.post(values.LOGIN_REQ_URL, headers=values.HEADERS, data=login_payload)
# #TODO ensure result is correct and login was succesfull.



def formatDateString(unformatedString: string) -> string:
    return unformatedString.split("<span>")


rota_html = s.get(values.ROTA_URL)
soup = bs(rota_html.text, "html.parser")

pprint(soup.prettify())

"""Deal with the dates and title"""
dates = soup.findAll("th",{"class":"day"}, limit=7) #Collect the first 7 results (each day of the week)

wc_date = dates[0].get_text()[3:8]  #Used for file naming and title of spreadsheet, removes /n
day_strings = [day.get_text().strip()[:3] + " " + day.get_text().strip()[3:]    #Adds a space and cleans up
               for day in dates]



"""" Deals with the employees and their rotas  """

rotas = soup.findAll(attrs={"class": "employee-name"},
                     string=values.KITCHEN_NAMES)  # You can pass a list in? AWESOME!


# TODO For each kitchen staff, jump to parent element then extract all class='day text-center"
extracted_rotas = {}
for manager in rotas:
    shifts = []
    managerName = manager.text
    week = manager.find_parent().findAll(attrs={"class": "day"})
    for day in week:
        strippedText = day.get_text("<br>").strip()  # Strips leading escape chars . Value will be
        # similar to '06:00am - 03:00pm                Kitchen)'

        # TODO figure out how to cope with split shifts . if len(splitText) > 4 indicates muktiple
        #This just strips out any split shifts for now
        #Could use [::2] to just extract times as that's how it's formatted
        #len% 4 could show how many shifts, and each 4th entry is the shift type

        splitText = strippedText.split()[:4]  # Now looks something like
                                        # ['12:00pm', '-', '10:00pm', '<br>(Kitchen)<br>']

        if(len(strippedText) != 0):

            # convert times to 24 hour clock
            # %H is the 24 hour clock, %I is the 12 hour clock and when using the 12 hour clock,
            # %p qualifies if it is AM or PM. strp = parse, strf = format
            startTime = datetime.strptime(splitText[0], "%I:%M%p")  #This will have year and date attached
            startTime = datetime.strftime(startTime,"%H:%M")    #Strips that out, is now 24 hour 14:45
            endTime = datetime.strptime(splitText[2], "%I:%M%p")
            endTime = datetime.strftime(endTime, "%H:%M")

            # TODO ignore meetings


            shifts.append({"start": startTime,
                           "end":   endTime})
            #Example data structure
            # 'ROBBINS, WILL': [None,
            #                   {'end': '20:30', 'start': '11:00'},
            #                   {'end': '23:45', 'start': '17:30'},

        else: #If no shift that day
            shifts.append({"start": "day",
                         "end":    "off"})


        extracted_rotas[managerName] = shifts


wb = openpyxl.load_workbook(filename="compliance-tracker-template3.xlsx")
ws = wb["Kitchen LC Compliance"]      #wb workbook ws worksheet

"""Populate the spreadsheet template"""
ws["A1"] = f"Kitchen Line Check Compliance : w/c {wc_date}"
for row_index, manager in enumerate(extracted_rotas, values.NAME_COLUMN_START):

    ws[f"A{row_index}"] = manager


    day_colums = string.ascii_uppercase[values.MAPPINGS['MONDAY']:15]  #6 kitchen staff #B - H for example
    for x, day in enumerate(day_colums[::2]):    #Add days of week and dates,
        print(ws[f"{day}3"])
        ws[f"{day}3"] = day_strings[x]



    #TODO allow for dynamic amount of staff above

    for y, day in enumerate(day_colums[::2]):
        #TODO make this more readable - use named variables for start/finish times. Memory is not an issue
        ws[f"{day}{row_index}"] = (f"{extracted_rotas[manager][y]['start']}"
                           f"-"
                           f"{extracted_rotas[manager][y]['end']}"
                                   )
        if extracted_rotas[manager][y]['start'] != "day":
            #TODO THIS IS REALLY SLOPPY
            ws[f"{day_colums[(y*2) + 1]}{row_index}"] = "N"

                          #COULD USE INDEX OF day in colums

filename = f"compliance-tracker-wc{wc_date.replace('/', '-')}.xlsx"
print(filename)
wb.save(filename=filename)

"""EMAIL THE COMPLIANCE TRACKER"""

email = MIMEMultipart()
email["From"] = values.EMAIL_LOGIN
email["To"] = values.PUB_EMAIL
email["Subject"] = f"Kitchen LC Compliance tracker {wc_date}"
email.attach(MIMEText("Auto generated template for this week", "plain"))
with open(filename, "rb") as attachment:  #Read binary

    payload = MIMEBase('application', "octet-stream") #The "octet-stream" subtype is used to indicate
                                                        # that a body contains arbitrary binary data.
    payload.set_payload((attachment.read()))
    encoders.encode_base64(payload) #Must be base64 encoded to send
    payload.add_header("Content-Disposition",
                        "attachment; filename={}".format(Path(filename).name))
    email.attach(payload)

    with smtplib.SMTP("smtp.gmail.com") as connection:
        connection.starttls()
        connection.login(user=values.EMAIL_LOGIN, password=values.EMAIL_APP_PASSWORD)
        connection.sendmail(
            from_addr=values.EMAIL_LOGIN,
            to_addrs=values.PUB_EMAIL,
            msg=email.as_string()
        )





